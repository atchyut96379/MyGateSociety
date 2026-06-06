import io
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..constants import COMMITTEE_ROLE_LABELS, COMMITTEE_ROLES
from ..database import get_db
from ..excel_import import (
    _find_header_row,
    build_column_map,
    parse_excel_row,
    parse_phone_value,
)
from ..models import User
from ..schemas import (
    BulkImportResponse,
    BulkImportRowResult,
    CreateUserRequest,
    CreateUserResponse,
    CredentialsResponse,
    UpdateUserLoginRequest,
    UserResponse,
)
from ..security import get_current_user, require_secretary_ready
from ..user_service import (
    create_user_record,
    resolve_flat_id,
    update_user_login,
    update_user_role,
)

router = APIRouter(prefix="/users", tags=["users"])


def user_response(user: User) -> UserResponse:
    return UserResponse(
        id=user.id,
        email=user.email,
        phone=user.phone,
        name=user.name,
        role=user.role,
        society_id=user.society_id,
        flat_id=user.flat_id,
        flat_label=user.flat.label if user.flat else None,
        resident_type=user.resident_type,
        committee_role=user.committee_role,
        tenant_owner_name=user.tenant_owner_name,
        tenant_owner_phone=user.tenant_owner_phone,
        tenant_owner_flat_label=user.tenant_owner_flat_label,
        is_main_admin=user.is_main_admin,
        must_change_password=user.must_change_password,
    )


def credentials_from_user(user: User, plain_password: str, flat_label: str | None) -> CredentialsResponse:
    return CredentialsResponse(
        name=user.name,
        phone=user.phone,
        password=plain_password,
        role=user.role,
        resident_type=user.resident_type,
        committee_role=user.committee_role,
        flat_label=flat_label,
        must_change_password=True,
    )


@router.get("", response_model=list[UserResponse])
def list_users(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"ADMIN", "COMMITTEE"}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    users = (
        db.query(User)
        .filter(User.society_id == current_user.society_id)
        .order_by(User.created_at.desc())
        .all()
    )
    return [user_response(user) for user in users]


@router.post("", response_model=CreateUserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    payload: CreateUserRequest,
    admin: User = Depends(require_secretary_ready),
    db: Session = Depends(get_db),
):
    user, plain_password, flat_label = create_user_record(db, admin, payload)
    db.commit()
    db.refresh(user)
    return CreateUserResponse(
        user=user_response(user),
        credentials=credentials_from_user(user, plain_password, flat_label),
    )


def _parse_import_row(
    parsed: dict[str, Any],
    db: Session,
    admin: User,
) -> BulkImportRowResult:
    row_num = parsed["row"]
    if "error" in parsed:
        return BulkImportRowResult(
            row=row_num,
            name=parsed.get("name", "—"),
            phone=parsed.get("phone", ""),
            ok=False,
            error=parsed["error"],
        )

    name = parsed["name"]
    flat_label = parsed["flat_label"]
    role = parsed["role"]
    resident_type = parsed["resident_type"]
    committee_role = parsed.get("committee_role")
    email = parsed.get("email")

    try:
        phone = parse_phone_value(parsed["phone_raw"], flat_label, db, admin.society_id)
        flat_id = resolve_flat_id(db, admin.society_id, flat_label) if flat_label else None
        if role in {"RESIDENT", "COMMITTEE"} and not flat_id:
            raise HTTPException(status_code=400, detail="Flat required")

        payload = CreateUserRequest(
            name=name,
            phone=phone,
            email=email,
            role=role,
            resident_type=resident_type if role in {"RESIDENT", "COMMITTEE"} else None,
            committee_role=committee_role if role == "COMMITTEE" else None,
            flat_id=flat_id,
        )
        user, plain_password, resolved_flat = create_user_record(
            db, admin, payload, import_mode=True
        )
        db.flush()
        return BulkImportRowResult(
            row=row_num,
            name=user.name,
            phone=user.phone,
            ok=True,
            password=plain_password,
            flat_label=resolved_flat,
            role=user.role,
        )
    except HTTPException as exc:
        return BulkImportRowResult(
            row=row_num, name=name, phone=parsed.get("phone_raw", ""), ok=False, error=str(exc.detail)
        )
    except Exception as exc:
        return BulkImportRowResult(
            row=row_num, name=name, phone=parsed.get("phone_raw", ""), ok=False, error=str(exc)
        )


@router.get("/bulk-template")
@router.get("/import-template")
def download_import_template(_: User = Depends(require_secretary_ready)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["name", "phone", "flat", "Owner", "Tenant", "committee_role", "email"])
    ws.append(["Rama Rao", "9876543210", "119", "Owner", "", "", ""])
    ws.append(["Tenant Example", "", "120", "", "Tenant", "", ""])
    ws.append(["Duplex Owner", "9494974697 / 9493308460", "109/110", "Owner", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user-import-template.xlsx"},
    )


async def _import_users_from_excel(
    file: UploadFile,
    admin: User,
    db: Session,
) -> BulkImportResponse:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload an Excel file (.xlsx)")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    header_idx = _find_header_row(rows)  # noqa: SLF001
    mapping = build_column_map(rows[header_idx])

    if "name" not in mapping["fields"]:
        found = [str(c).strip() for c in rows[header_idx] if c is not None and str(c).strip()]
        raise HTTPException(
            status_code=400,
            detail=(
                "Excel must have a 'name' column (e.g. name, resident, member). "
                f"Found headers: {', '.join(found) or 'none'}. "
                "Use Download template on this page."
            ),
        )

    results: list[BulkImportRowResult] = []
    created = 0

    for row_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(row):
            continue
        parsed = parse_excel_row(row, row_num, mapping)
        if parsed is None:
            continue
        result = _parse_import_row(parsed, db, admin)
        results.append(result)
        if result.ok:
            created += 1

    db.commit()
    failed = len(results) - created
    return BulkImportResponse(created=created, failed=failed, results=results)


@router.post("/bulk-upload", response_model=BulkImportResponse)
async def bulk_upload_users(
    file: UploadFile = File(...),
    admin: User = Depends(require_secretary_ready),
    db: Session = Depends(get_db),
):
    return await _import_users_from_excel(file, admin, db)


@router.post("/import", response_model=BulkImportResponse, include_in_schema=False)
async def import_users(
    file: UploadFile = File(...),
    admin: User = Depends(require_secretary_ready),
    db: Session = Depends(get_db),
):
    return await _import_users_from_excel(file, admin, db)


@router.get("/committee-roles")
def committee_roles(_: User = Depends(get_current_user)):
    return [
        {"value": r, "label": COMMITTEE_ROLE_LABELS.get(r, r.replace("_", " ").title())}
        for r in COMMITTEE_ROLES
    ]


@router.get("/{user_id}", response_model=UserResponse)
def get_user(
    user_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in {"ADMIN", "COMMITTEE"}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    user = db.get(User, user_id)
    if user is None or user.society_id != current_user.society_id:
        raise HTTPException(status_code=404, detail="User not found")
    return user_response(user)


@router.patch("/{user_id}/role", response_model=UserResponse)
def update_user_role_endpoint(
    user_id: str,
    payload: UpdateUserLoginRequest,
    admin: User = Depends(require_secretary_ready),
    db: Session = Depends(get_db),
):
    user = update_user_role(db, admin, user_id, payload)
    db.commit()
    db.refresh(user)
    return user_response(user)


@router.patch("/{user_id}/login", response_model=CreateUserResponse)
def update_user_login_endpoint(
    user_id: str,
    payload: UpdateUserLoginRequest,
    admin: User = Depends(require_secretary_ready),
    db: Session = Depends(get_db),
):
    user, plain_password, flat_label = update_user_login(db, admin, user_id, payload)
    db.commit()
    db.refresh(user)
    return CreateUserResponse(
        user=user_response(user),
        credentials=credentials_from_user(user, plain_password, flat_label),
    )
